#!/usr/bin/env python3
"""
Rebuilds the "Score Sheet" tab and adds a "Wine Log" tab in the Tasting Night
prediction workbook, in place. Leaves Lineup, Predictions, and _pred_lookup
(and their formulas) untouched.

Protocol (Tasting Night 3+): tastings are labels-visible, paper on the night,
bulk-uploaded after via ingest-tasting-sheet.mjs. Each taster rates BOTH wines
of a pairing on the app's 1-100 scale, picks a winner (ties allowed), and
gives a short "why" (free text). Every bottle is logged with price; ringer
bottles have price hidden from tasters but recorded in the log.

Score Sheet columns (A-J):
  Taster | Pairing (P1-P6/free) | Wine A # | Wine B # | Rating A (1-100) |
  Rating B (1-100) | Winner (# or 'tie') | Why (few words) | Predicted pick |
  Correct?

"Predicted pick" and "Correct?" keep the workbook's existing INDEX/MATCH
formulas against the hidden _pred_lookup tab (key = Taster|Pairing), just
re-pointed at the new column positions (Winner moved from E to G).

Wine Log columns (A-J):
  # | Producer | Name | Vintage | Wine Type | Country | Region | Grapes |
  Price | Ringer?

Pre-filled from the Lineup tab's 14 archetypes: the "#" column and (for
archetype #12, the deliberate blind-QPR ringer) "Ringer?" = yes. Everything
bottle-specific (producer, name, vintage, price, ...) is left blank to fill
in on purchase. Each "#" cell carries a comment with the archetype name as a
purchasing reminder.

Usage:
  python3 scripts/somm-eval/build-tasting-score-sheet.py [--path=/path/to/workbook.xlsx]

Defaults to the Tasting Night 3 workbook in ~/Downloads. Requires openpyxl
(pip install openpyxl).
"""
import argparse
import copy
import os
import sys

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_PATH = os.path.expanduser(
    "~/Downloads/Cluster_Tasting3_Predictions.xlsx"
)

DATA_ROWS = 120  # matches the sheet's existing row count (~120 rows)

# Grenache brand styling, lifted from the workbook's existing Score Sheet header.
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(patternType="solid", fgColor="7B1D3A")
HEADER_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER_SIDE = Side(style="thin", color="E8E0D0")
DATA_BORDER = Border(
    left=THIN_BORDER_SIDE,
    right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,
    bottom=THIN_BORDER_SIDE,
)
TIP_FONT = Font(name="Arial", size=9, italic=True)

SCORE_SHEET_HEADERS = [
    "Taster",
    "Pairing (P1-P6 or free)",
    "Wine A #",
    "Wine B #",
    "Rating A (1-100)",
    "Rating B (1-100)",
    "Winner (# or 'tie')",
    "Why (few words)",
    "Predicted pick",
    "Correct?",
]
SCORE_SHEET_WIDTHS = [12, 22, 9, 9, 13, 13, 16, 34, 16, 10]

SCORE_SHEET_PROTOCOL_NOTE = (
    "Protocol: RATE BOTH wines in every pairing on your usual 1-100 scale "
    "(Rating A, Rating B). Winner = the wine # that won, or write 'tie' "
    "— ties are allowed and should be recorded honestly, not forced. "
    "Why = a few plain words on what tipped it, not a paragraph. LOG EVERY "
    "WINE poured in the Wine Log tab, including ringers (their price is "
    "hidden from tasters but must still be recorded). 'Predicted pick' "
    "fills automatically when Taster + Pairing match the Predictions tab."
)

WINE_LOG_HEADERS = [
    "#",
    "Producer",
    "Name",
    "Vintage",
    "Wine Type",
    "Country",
    "Region",
    "Grapes",
    "Price",
    "Ringer?",
]
WINE_LOG_WIDTHS = [6, 22, 26, 10, 12, 14, 18, 20, 10, 10]

# Archetype # that's the deliberate blind-QPR ringer (see Lineup tab row 13).
RINGER_ARCHETYPE_NUMBER = "12"


def style_header_row(ws, num_cols, headers, widths):
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


def style_data_cell(cell):
    cell.font = copy.copy(DATA_FONT)
    cell.border = copy.copy(DATA_BORDER)


def read_lineup_archetypes(wb):
    """Returns [{"number": "1", "archetype": "...", ...}, ...] for all 14 rows."""
    ws = wb["Lineup"]
    archetypes = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        number, archetype = row[0], row[1]
        if number is None:
            continue
        archetypes.append({"number": str(number), "archetype": archetype})
    return archetypes


def rebuild_score_sheet(wb):
    # Recreate in place so it keeps its position in the tab order.
    old_index = wb.sheetnames.index("Score Sheet")
    del wb["Score Sheet"]
    ws = wb.create_sheet("Score Sheet", old_index)

    style_header_row(ws, len(SCORE_SHEET_HEADERS), SCORE_SHEET_HEADERS, SCORE_SHEET_WIDTHS)

    last_data_row = 1 + DATA_ROWS
    for r in range(2, last_data_row + 1):
        for c in range(1, len(SCORE_SHEET_HEADERS) + 1):
            style_data_cell(ws.cell(row=r, column=c))
        # I: Predicted pick — unchanged lookup, Taster (A) & Pairing (B) columns didn't move.
        ws.cell(
            row=r,
            column=9,
            value=(
                f'=IFERROR(INDEX(_pred_lookup!B:B,MATCH(A{r}&"|"&B{r},'
                f'_pred_lookup!A:A,0)),"")'
            ),
        )
        # J: Correct? — re-pointed at Winner (now G) and Predicted pick (now I).
        ws.cell(
            row=r,
            column=10,
            value=(
                f'=IF(OR(I{r}="",G{r}=""),"",IF(G{r}=IFERROR(INDEX(_pred_lookup!C:C,'
                f'MATCH(A{r}&"|"&B{r},_pred_lookup!A:A,0)),-1),"YES","no"))'
            ),
        )

    note_row = last_data_row + 2  # one blank row, matching the original layout
    ws.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=len(SCORE_SHEET_HEADERS)
    )
    note_cell = ws.cell(row=note_row, column=1, value=SCORE_SHEET_PROTOCOL_NOTE)
    note_cell.font = TIP_FONT
    note_cell.alignment = Alignment(wrap_text=True)

    return ws


def build_wine_log(wb):
    if "Wine Log" in wb.sheetnames:
        del wb["Wine Log"]
    # Insert right after Lineup: Lineup defines the archetypes, Wine Log
    # records the real bottles bought against them.
    lineup_index = wb.sheetnames.index("Lineup")
    ws = wb.create_sheet("Wine Log", lineup_index + 1)

    style_header_row(ws, len(WINE_LOG_HEADERS), WINE_LOG_HEADERS, WINE_LOG_WIDTHS)

    archetypes = read_lineup_archetypes(wb)
    for idx, entry in enumerate(archetypes):
        r = idx + 2
        for c in range(1, len(WINE_LOG_HEADERS) + 1):
            style_data_cell(ws.cell(row=r, column=c))
        number_cell = ws.cell(row=r, column=1, value=entry["number"])
        if entry.get("archetype"):
            number_cell.comment = Comment(str(entry["archetype"]), "Lineup tab")
        if entry["number"] == RINGER_ARCHETYPE_NUMBER:
            ws.cell(row=r, column=10, value="yes")

    return ws


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--path", default=DEFAULT_PATH, help="Path to the .xlsx workbook (edited in place)")
    args = parser.parse_args()

    if not os.path.exists(args.path):
        print(f"File not found: {args.path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.path)
    required = {"Lineup", "Predictions", "Score Sheet", "_pred_lookup"}
    missing = required - set(wb.sheetnames)
    if missing:
        print(f"Workbook is missing expected tabs: {sorted(missing)}", file=sys.stderr)
        sys.exit(1)

    rebuild_score_sheet(wb)
    build_wine_log(wb)

    wb.save(args.path)
    print(f"Saved {args.path}")

    # Verify: reload from disk and report.
    verify_wb = openpyxl.load_workbook(args.path)
    print("Tabs:", verify_wb.sheetnames)
    score_headers = [
        verify_wb["Score Sheet"].cell(row=1, column=c).value
        for c in range(1, len(SCORE_SHEET_HEADERS) + 1)
    ]
    print("Score Sheet headers:", score_headers)
    wine_log_headers = [
        verify_wb["Wine Log"].cell(row=1, column=c).value
        for c in range(1, len(WINE_LOG_HEADERS) + 1)
    ]
    print("Wine Log headers:", wine_log_headers)


if __name__ == "__main__":
    main()
